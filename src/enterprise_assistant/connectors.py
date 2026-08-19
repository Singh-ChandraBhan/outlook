from __future__ import annotations

from datetime import datetime, timedelta, timezone
from html import unescape
import re
from os import PathLike
from typing import Any, BinaryIO

import requests

from .config import Settings
from .models import KnowledgeRecord

DEMO_NOW = datetime(2025, 1, 15, 12, 0, tzinfo=timezone.utc)


def _strip_html(value: str) -> str:
    return unescape(re.sub(r"<[^>]+>", " ", value or "")).strip()


def demo_records() -> list[KnowledgeRecord]:
    stamp = DEMO_NOW.isoformat()
    specs = [
        (
            "outlook",
            "mail-001",
            "Quarterly architecture review",
            "Review is Friday at 10:00 UTC.",
            "email",
            "",
            "",
            "Asha",
            ["employees"],
        ),
        (
            "outlook",
            "event-001",
            "Architecture review meeting",
            "Teams meeting for platform decisions.",
            "calendar",
            "confirmed",
            "",
            "Asha",
            ["employees"],
        ),
        (
            "teams",
            "team-001",
            "Search rollout",
            "Hybrid search rollout completed successfully.",
            "message",
            "",
            "",
            "DevOps",
            ["employees", "engineering"],
        ),
        (
            "servicenow",
            "INC0010001",
            "Checkout API unavailable",
            "Production checkout requests fail. SLA remaining: 3 hours.",
            "incident",
            "open",
            "P1",
            "SRE",
            ["engineering"],
        ),
    ]
    return [
        KnowledgeRecord.create(
            s,
            sid,
            title,
            content,
            created_at=stamp,
            updated_at=stamp,
            record_type=typ,
            status=status,
            priority=priority,
            owner=owner,
            allowed_groups=groups,
            metadata={"sla_hours": 3} if sid == "INC0010001" else {},
        )
        for s, sid, title, content, typ, status, priority, owner, groups in specs
    ]


class GraphConnector:
    BASE = "https://graph.microsoft.com/v1.0"

    def __init__(self, settings: Settings):
        self.settings = settings

    def _token(self) -> str:
        from azure.identity import ClientSecretCredential

        credential = ClientSecretCredential(
            self.settings.graph_tenant_id,
            self.settings.graph_client_id,
            self.settings.graph_client_secret,
        )
        return credential.get_token("https://graph.microsoft.com/.default").token

    def _get(
        self, path: str, params: dict[str, Any] | None = None
    ) -> list[dict[str, Any]]:
        response = requests.get(
            f"{self.BASE}{path}",
            params=params,
            headers={"Authorization": f"Bearer {self._token()}"},
            timeout=30,
        )
        response.raise_for_status()
        return response.json().get("value", [])

    def messages(self) -> list[KnowledgeRecord]:
        rows = self._get(
            f"/users/{self.settings.graph_user_id}/messages",
            {
                "$top": 50,
                "$select": "id,subject,bodyPreview,receivedDateTime,lastModifiedDateTime,webLink,from",
            },
        )
        return [
            KnowledgeRecord.create(
                "outlook",
                x["id"],
                x.get("subject") or "(no subject)",
                x.get("bodyPreview", ""),
                created_at=x.get("receivedDateTime", ""),
                updated_at=x.get("lastModifiedDateTime", ""),
                url=x.get("webLink", ""),
                record_type="email",
                owner=x.get("from", {}).get("emailAddress", {}).get("address", ""),
            )
            for x in rows
        ]

    def calendar(self) -> list[KnowledgeRecord]:
        now = datetime.now(timezone.utc)
        rows = self._get(
            f"/users/{self.settings.graph_user_id}/calendarView",
            {
                "startDateTime": (
                    now - timedelta(days=self.settings.calendar_days_past)
                ).isoformat(),
                "endDateTime": (
                    now + timedelta(days=self.settings.calendar_days_future)
                ).isoformat(),
                "$top": 50,
            },
        )
        return [
            KnowledgeRecord.create(
                "outlook",
                x["id"],
                x.get("subject") or "Meeting",
                _strip_html(x.get("body", {}).get("content", "")),
                created_at=x.get("createdDateTime", ""),
                updated_at=x.get("lastModifiedDateTime", ""),
                url=x.get("webLink", ""),
                record_type="calendar",
                status="cancelled" if x.get("isCancelled") else "confirmed",
                metadata={"start": x.get("start"), "end": x.get("end")},
            )
            for x in rows
        ]

    def teams(self) -> list[KnowledgeRecord]:
        rows = self._get(
            f"/teams/{self.settings.graph_team_id}/channels/{self.settings.graph_channel_id}/messages",
            {"$top": 50},
        )
        return [
            KnowledgeRecord.create(
                "teams",
                x["id"],
                x.get("subject") or "Teams message",
                _strip_html(x.get("body", {}).get("content", "")),
                created_at=x.get("createdDateTime", ""),
                updated_at=x.get("lastModifiedDateTime")
                or x.get("createdDateTime", ""),
                url=x.get("webUrl", ""),
                record_type="message",
                owner=x.get("from", {}).get("user", {}).get("displayName", ""),
            )
            for x in rows[:50]
        ]


def outlook_desktop_calendar(settings: Settings) -> list[KnowledgeRecord]:
    if not settings.outlook_desktop_mailbox:
        return []
    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("Install pywin32 to use Outlook Desktop") from exc
    namespace = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    recipient = namespace.CreateRecipient(settings.outlook_desktop_mailbox)
    recipient.Resolve()
    if not recipient.Resolved:
        raise RuntimeError("OUTLOOK_DESKTOP_MAILBOX could not be resolved")
    mailbox = settings.outlook_desktop_mailbox.casefold()
    signed_in = {
        str(account.SmtpAddress).casefold()
        for account in namespace.Accounts
        if getattr(account, "SmtpAddress", None)
    }
    # GetSharedDefaultFolder can fail for the user's own mailbox. Outlook's
    # default folder API is the correct path for the signed-in account.
    folder = (
        namespace.GetDefaultFolder(9)
        if mailbox in signed_in
        else namespace.GetSharedDefaultFolder(recipient, 9)
    )
    items = folder.Items
    items.Sort("[Start]")
    items.IncludeRecurrences = True
    now = datetime.now().astimezone().replace(tzinfo=None)
    window_start = now - timedelta(days=settings.calendar_days_past)
    window_end = now + timedelta(days=settings.calendar_days_future)
    records = []
    for item in items:
        source_id = getattr(item, "EntryID", "")
        start = getattr(item, "Start", None)
        start_value = start.replace(tzinfo=None) if start else None
        if source_id and start_value and window_start <= start_value <= window_end:
            organizer = str(getattr(item, "Organizer", "") or "")
            location = str(getattr(item, "Location", "") or "")
            required = str(getattr(item, "RequiredAttendees", "") or "")
            optional = str(getattr(item, "OptionalAttendees", "") or "")
            join_url = str(getattr(item, "NetMeetingUrl", "") or "")
            records.append(
                KnowledgeRecord.create(
                    "outlook-desktop",
                    source_id,
                    str(getattr(item, "Subject", "") or "(no subject)"),
                    str(getattr(item, "Body", "") or ""),
                    record_type="calendar",
                    status=str(getattr(item, "MeetingStatus", "") or ""),
                    owner=organizer,
                    metadata={
                        "start": str(start),
                        "end": str(getattr(item, "End", "") or ""),
                        "organizer": organizer,
                        "location": location,
                        "attendees": [
                            x.strip() for x in required.split(";") if x.strip()
                        ],
                        "optional_attendees": [
                            x.strip() for x in optional.split(";") if x.strip()
                        ],
                        "join_url": join_url,
                    },
                )
            )
    return records


class ServiceNowConnector:
    def __init__(self, settings: Settings):
        self.settings = settings

    def incidents(self) -> list[KnowledgeRecord]:
        url = f"{self.settings.servicenow_url.rstrip('/')}/api/now/table/{self.settings.servicenow_table}"
        response = requests.get(
            url,
            auth=(self.settings.servicenow_username, self.settings.servicenow_password),
            params={"sysparm_limit": 100, "sysparm_display_value": "true"},
            timeout=30,
        )
        response.raise_for_status()
        return [self._normalize(row) for row in response.json().get("result", [])]

    @staticmethod
    def _normalize(row: dict[str, Any]) -> KnowledgeRecord:
        sid = str(row.get("number") or row.get("sys_id"))
        return KnowledgeRecord.create(
            "servicenow",
            sid,
            str(row.get("short_description", "Incident")),
            str(row.get("description", "")),
            created_at=str(row.get("sys_created_on", "")),
            updated_at=str(row.get("sys_updated_on", "")),
            record_type="incident",
            status=str(row.get("state", "")),
            priority=str(row.get("priority", "")),
            owner=str(row.get("assigned_to", "")),
            metadata={"sla_hours": row.get("sla_hours")},
        )


def incidents_from_xlsx(path: str | PathLike[str] | BinaryIO) -> list[KnowledgeRecord]:
    from openpyxl import load_workbook

    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(x).strip() for x in next(rows)]
    return [
        ServiceNowConnector._normalize(dict(zip(headers, values)))
        for values in rows
        if any(values)
    ]


class ExcelIncidentConnector:
    """Compatibility connector used by the Streamlit dashboard."""

    def __init__(self, workbook: str | PathLike[str] | BinaryIO):
        self.workbook = workbook

    def collect(self) -> list[KnowledgeRecord]:
        return incidents_from_xlsx(self.workbook)


class OutlookDesktopConnector:
    """Read calendar records from an already signed-in classic Outlook profile."""

    def __init__(self, mailbox: str, days_past: int = 7, days_future: int = 30):
        if not mailbox.strip():
            raise ValueError("Enter the mailbox configured in classic Outlook")
        self.settings = Settings(
            outlook_desktop_mailbox=mailbox.strip(),
            calendar_days_past=days_past,
            calendar_days_future=days_future,
        )

    def collect(self) -> list[KnowledgeRecord]:
        records = outlook_desktop_calendar(self.settings)
        for record in records:
            record.record_type = "calendar_event"
            record.metadata = {
                "meeting_start": record.metadata.get("start", ""),
                "meeting_end": record.metadata.get("end", ""),
                "organizer": record.owner,
                **record.metadata,
            }
        return records
